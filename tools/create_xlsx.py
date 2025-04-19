#!/usr/bin/env python

import os, sys

file_path = 'tools/python/'
sys.path.append(os.path.dirname(file_path))

import csv
import openpyxl

export_file = sys.argv[1]

file_report = openpyxl.Workbook()


foglio0 = file_report.active
foglio0.title = "main"
foglio0["A1"] = "Sheet Name"
foglio0["A2"] = "d_ftdd"
foglio0["A3"] = "d_nr"
foglio0["A4"] = "d_cells"
foglio0["A5"] = "d_al"
foglio0["A6"] = "d_al_pt"
foglio0["A7"] = "d_al_mo"
foglio0["A8"] = "d_upg"
foglio0["A9"] = "ftdd"
foglio0["A10"] = "nr"
foglio0["A11"] = "cells"
foglio0["A12"] = "cu"
foglio0["A13"] = "sleep"
foglio0["A14"] = "al"
foglio0["A15"] = "up"
foglio0["A16"] = "cv"


foglio0["B1"] = "Sheet Description"
foglio0["B2"]= "The current and the previous status of EUtranCellFDD, EUtranCellTDD. It includes the cell sleep status"
foglio0["B3"]= "The current and the previous status of NRCellDU and NRCellCU"
foglio0["B4"]= "The current and the previous status of all the Cells (Trx, EUtranCellFDD, EUtranCellTDD, NRCellDu)"
foglio0["B5"]= "The current alarms except the ones present before the activity"
foglio0["B6"]= "The current alarms except the ones present before the activity and grouped by problem type"
foglio0["B7"]= "The current alarms except the ones present before the activity and grouped by MO"
foglio0["B8"]= "The current and the previous upgrade package executing on the node. The field is filled only if the upgrade package hasn't changed"
foglio0["B9"]= "The current EUtranCellFDD, EUtranCellTDD and Sleep Configuration"
foglio0["B10"]= "The current NRCell CU and DU status"
foglio0["B11"]= "The current status of all the cells"
foglio0["B12"]= "The current status of the NRCellCU"
foglio0["B13"]= "The current status of the Sleep"
foglio0["B14"]= "The current alarms"
foglio0["B15"]= "The current upgrade package executing on the node"
foglio0["B16"]= "The last CV"

foglio0.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 10
foglio0.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 90

foglio0.sheet_properties.tabColor = "FF0000"
file_report.save(export_file)


