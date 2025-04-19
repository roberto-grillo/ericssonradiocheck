#!/usr/bin/env python

import os, sys
file_path = 'tools/python/'
sys.path.append(os.path.dirname(file_path))

import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.comments import Comment

export_file = sys.argv[1]
input_file1 = sys.argv[2]
input_file2 = sys.argv[3]
input_file3 = sys.argv[4]
input_file4 = sys.argv[5]
input_file5 = sys.argv[6]
input_file6 = sys.argv[7]
input_file7 = sys.argv[8]
input_file8 = sys.argv[9]
input_file9 = sys.argv[10]
input_file10 = sys.argv[11]
input_file11 = sys.argv[12]
input_file12 = sys.argv[13]
input_file13 = sys.argv[14]
input_file14 = sys.argv[15]
input_file15 = sys.argv[16]
input_file16 = sys.argv[17]
input_file17 = sys.argv[18]
input_file18 = sys.argv[19]


print("Exporting csv file:", input_file1)
print("Exporting csv file:", input_file2)
print("Exporting csv file:", input_file3)
print("Exporting csv file:", input_file4)
print("Exporting csv file:", input_file5)
print("Exporting csv file:", input_file6)
print("Exporting csv file:", input_file7)
print("Exporting csv file:", input_file8)
print("Exporting csv file:", input_file9)
print("Exporting csv file:", input_file10)
print("Exporting csv file:", input_file11)
print("Exporting csv file:", input_file12)
print("Exporting csv file:", input_file13)
print("Exporting csv file:", input_file14)
print("Exporting csv file:", input_file15)
print("Exporting csv file:", input_file16)
print("Exporting csv file:", input_file17)
print("Exporting csv file:", input_file18)
print("Creating xlsx file:", export_file)

file_report = openpyxl.load_workbook(export_file)

#file_report = openpyxl.Workbook()
#foglio0 = file_report.active
#foglio0.title = "main"
#foglio0["A1"] = "Double Report"

border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

foglio1 = file_report.create_sheet("d_ftdd")
with open(input_file1) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio1.append(row)
#last_column = foglio1.max_column
#last_row = foglio1.max_row

foglio2 = file_report.create_sheet("d_nr")
with open(input_file2) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio2.append(row)

foglio3 = file_report.create_sheet("d_cells")
with open(input_file3) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio3.append(row)

foglio4 = file_report.create_sheet("d_al")
with open(input_file4) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio4.append(row)

foglio5 = file_report.create_sheet("d_al_pt")
with open(input_file5) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio5.append(row)

foglio6 = file_report.create_sheet("d_al_mo")
with open(input_file6) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio6.append(row)

foglio7 = file_report.create_sheet("d_upg")
with open(input_file7) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio7.append(row)



foglio8 = file_report.create_sheet("ftdd")
with open(input_file8) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio8.append(row)

foglio9 = file_report.create_sheet("nr")
with open(input_file9) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio9.append(row)

foglio10 = file_report.create_sheet("cells")
with open(input_file10) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio10.append(row)

foglio11 = file_report.create_sheet("cu")
with open(input_file11) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio11.append(row)

foglio12 = file_report.create_sheet("sleep")
with open(input_file12) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio12.append(row)

foglio13 = file_report.create_sheet("al")
with open(input_file13) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio13.append(row)

foglio14 = file_report.create_sheet("up")
with open(input_file14) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio14.append(row)

foglio15 = file_report.create_sheet("cv")
with open(input_file15) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio15.append(row)

foglio16 = file_report.create_sheet("rn_lic")
with open(input_file16) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio16.append(row)

foglio17 = file_report.create_sheet("rn_em")
with open(input_file17) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio17.append(row)

foglio18 = file_report.create_sheet("erbs_lic")
with open(input_file18) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio18.append(row)


foglio1.sheet_properties.tabColor = "FFA500"
foglio2.sheet_properties.tabColor = "FFA500"
foglio3.sheet_properties.tabColor = "FFA500"
foglio4.sheet_properties.tabColor = "FFA500"
foglio5.sheet_properties.tabColor = "FFA500"
foglio6.sheet_properties.tabColor = "FFA500"
foglio7.sheet_properties.tabColor = "FFA500"
foglio8.sheet_properties.tabColor = "007FFF"
foglio9.sheet_properties.tabColor = "007FFF"
foglio10.sheet_properties.tabColor = "007FFF"
foglio11.sheet_properties.tabColor = "007FFF"
foglio12.sheet_properties.tabColor = "007FFF"
foglio13.sheet_properties.tabColor = "007FFF"
foglio14.sheet_properties.tabColor = "007FFF"
foglio15.sheet_properties.tabColor = "007FFF"
foglio16.sheet_properties.tabColor = "007FFF"
foglio17.sheet_properties.tabColor = "007FFF"
foglio18.sheet_properties.tabColor = "007FFF"


colora_rosso = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
colora_verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
colora_aranc = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
colora_giall = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
colora_blu = PatternFill(start_color='007FFF', end_color='00FF00', fill_type='solid')
#cella_da_colorare = foglio9['A1']
#cella_da_colorare.fill = colora_verde

comment_text_pre = 'Data of Pre Activity'
comment_text_post = 'Data of Post Activity'
comment_author = 'RG'

comment_pre = Comment(comment_text_pre, comment_author)
comment_post = Comment(comment_text_post, comment_author)

font_size = 13
width = len(comment_text_post) * font_size / 2 # Width based on text length
height = 30
comment_pre.width = width
comment_post.width = width
comment_pre.height = height
comment_post.height = height

cella_da_colorare = foglio1['A1']
cella_da_colorare.fill = colora_giall

celle_da_colorare = foglio1['B1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_aranc
        cell.comment = comment_pre

celle_da_colorare = foglio1['G1:K1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu
        cell.comment = comment_post



cella_da_colorare = foglio2['A1']
cella_da_colorare.fill = colora_giall

celle_da_colorare = foglio2['B1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_aranc
        cell.comment = comment_pre

celle_da_colorare = foglio2['G1:K1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu
        cell.comment = comment_post


cella_da_colorare = foglio3['A1']
cella_da_colorare.fill = colora_giall

celle_da_colorare = foglio3['B1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_aranc
        cell.comment = comment_pre

celle_da_colorare = foglio3['E1:G1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu
        cell.comment = comment_post



celle_da_colorare = foglio4['A1:K1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio5['A1:C1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio6['A1:C1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


cella_da_colorare = foglio7['A1']
cella_da_colorare.fill = colora_giall

cella_da_colorare = foglio7['B1']
cella_da_colorare.fill = colora_aranc
cella_da_colorare.comment = comment_pre

cella_da_colorare = foglio7['C1']
cella_da_colorare.fill = colora_blu
cella_da_colorare.comment = comment_post



celle_da_colorare = foglio8['A1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio9['A1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio10['A1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio11['A1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio12['A1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


celle_da_colorare = foglio13['A1:K1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio14['A1:C1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio15['A1:C1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


celle_da_colorare = foglio16['A1:G1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


celle_da_colorare = foglio17['A1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu



#foglio3.sheet_state = "hidden"

for sheet_name in file_report.sheetnames:
    # Select the sheet by name
    sheet = file_report[sheet_name]

    # Autofit column widths for all columns
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Add filters to each filled column
    for column in sheet.iter_cols():
        if all(cell.value is None for cell in column):
            continue
        sheet.auto_filter.ref = sheet.dimensions

    # Add borders to every cell in each row
    #last_row = sheet.max_row
    #for row in sheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=15):
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border_style



file_report.save(export_file)
