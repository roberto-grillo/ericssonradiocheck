#!/usr/bin/env python

import os, sys
file_path = 'tools/python/'
sys.path.append(os.path.dirname(file_path))

import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

export_file = sys.argv[1]
input_file8 = sys.argv[9]
input_file9 = sys.argv[10]
input_file10 = sys.argv[11]
input_file11 = sys.argv[12]
input_file12 = sys.argv[13]
input_file13 = sys.argv[14]
input_file14 = sys.argv[15]
input_file15 = sys.argv[16]
input_file16 = sys.argv[17]
input_file17 = sys.argv[18]
input_file18 = sys.argv[19]



print("Exporting csv file:", input_file8)
print("Exporting csv file:", input_file9)
print("Exporting csv file:", input_file10)
print("Exporting csv file:", input_file11)
print("Exporting csv file:", input_file12)
print("Exporting csv file:", input_file13)
print("Exporting csv file:", input_file14)
print("Exporting csv file:", input_file15)
print("Exporting csv file:", input_file16)
print("Exporting csv file:", input_file17)
print("Exporting csv file:", input_file18)
print("Creating xlsx file:", export_file)

file_report = openpyxl.load_workbook(export_file)

#file_report = openpyxl.Workbook()
#foglio0 = file_report.active
#foglio0.title = "main"
#foglio0["A1"] = "Double Report"

border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


foglio8 = file_report.create_sheet("ftdd")
with open(input_file8) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio8.append(row)

foglio9 = file_report.create_sheet("nr")
with open(input_file9) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio9.append(row)

foglio10 = file_report.create_sheet("cells")
with open(input_file10) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio10.append(row)

foglio11 = file_report.create_sheet("cu")
with open(input_file11) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio11.append(row)

foglio12 = file_report.create_sheet("sleep")
with open(input_file12) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio12.append(row)

foglio13 = file_report.create_sheet("al")
with open(input_file13) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio13.append(row)

foglio14 = file_report.create_sheet("up")
with open(input_file14) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio14.append(row)

foglio15 = file_report.create_sheet("cv")
with open(input_file15) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio15.append(row)

foglio16 = file_report.create_sheet("rn_lic")
with open(input_file16) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio16.append(row)

foglio17 = file_report.create_sheet("rn_em")
with open(input_file17) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio17.append(row)

foglio18 = file_report.create_sheet("erbs_lic")
with open(input_file18) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        foglio18.append(row)


foglio8.sheet_properties.tabColor = "007FFF"
foglio9.sheet_properties.tabColor = "007FFF"
foglio10.sheet_properties.tabColor = "007FFF"
foglio11.sheet_properties.tabColor = "007FFF"
foglio12.sheet_properties.tabColor = "007FFF"
foglio13.sheet_properties.tabColor = "007FFF"
foglio14.sheet_properties.tabColor = "007FFF"
foglio15.sheet_properties.tabColor = "007FFF"
foglio16.sheet_properties.tabColor = "007FFF"
foglio17.sheet_properties.tabColor = "007FFF"
foglio18.sheet_properties.tabColor = "007FFF"

colora_rosso = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
colora_verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
colora_blu = PatternFill(start_color='007FFF', end_color='00FF00', fill_type='solid')
#cella_da_colorare = foglio9['A1']
#cella_da_colorare.fill = colora_verde


celle_da_colorare = foglio8['A1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio9['A1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio10['A1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio11['A1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio12['A1:D1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


celle_da_colorare = foglio13['A1:K1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio14['A1:C1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio15['A1:C1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


celle_da_colorare = foglio16['A1:G1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu

celle_da_colorare = foglio17['A1:F1']
for row in celle_da_colorare:
    for cell in row:
        cell.fill = colora_blu


for sheet_name in file_report.sheetnames:
    # Select the sheet by name
    sheet = file_report[sheet_name]

    # Autofit column widths for all columns
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Add filters to each filled column
    for column in sheet.iter_cols():
        if all(cell.value is None for cell in column):
            continue
        sheet.auto_filter.ref = sheet.dimensions

    # Add borders to every cell in each row
    #last_row = sheet.max_row
    #for row in sheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=15):
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border_style


file_report.save(export_file)
